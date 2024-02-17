import random
import time
from random import randint

import openpyxl
import pickle
import json
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
# from selenium import webdriver
#
# driver = webdriver.Chrome()

import undetected_chromedriver as uc
from time import sleep
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("user-data-dir=C:\\Users\\alienware\\AppData\\Local\\Google\\Chrome\\User Data")
chrome_options.add_argument("profile-directory=Profile 2")
driver = uc.Chrome()
driver.get("https://booking.corbettgov.org/resort/")

# cookies = pickle.load(open("cookies.pkl", "rb"))
#
# for cookie in cookies:
#     driver.add_cookie(cookie)
# driver.get("https://booking.corbettgov.org/resort/")
driver.maximize_window()

input("please press enter when you want to start putting details from excel sheet")

#get excel data below
workbook = openpyxl.load_workbook("paxData.xlsx")
sheet = workbook['pax']
rows = sheet.max_row
cols = sheet.max_column
print(rows)


secLoop = 1
loops=1

for r in range(2,rows+1):
    for c in range(1,cols+1):
        print("Entering... "+str(sheet.cell(r,c).value),end = " ")
        print("")

        try:
            # driver.find_element(By.XPATH,
            #                     "((//div[contains(@class,' font-monospace')])[1]//div[contains(@class,'form-floating')])[" + str(
            #                         loops) + "]/select").click()
            Select(driver.find_element(By.XPATH,"((//div[contains(@class,' font-monospace')])[1]//div[contains(@class,'form-floating')])["+str(loops)+"]/select")).select_by_visible_text(sheet.cell(r,c).value)
            sleep(random.uniform(5, 7))
        except:
            # driver.find_element(By.XPATH,
            #                     "((//div[contains(@class,' font-monospace')])[1]//div[contains(@class,'form-floating')])[" + str(
            #                         loops) + "]/input").click()
            driver.find_element(By.XPATH,
                                "((//div[contains(@class,' font-monospace')])[1]//div[contains(@class,'form-floating')])[" + str(loops) + "]/input").send_keys(
                sheet.cell(r, c).value)
            sleep(random.uniform(5, 7))


        loops += 1


driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")

print("all fields entered successfully")



input("please any key to exit program")
